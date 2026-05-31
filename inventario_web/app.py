from flask import Flask, Response, render_template, request, redirect, url_for, flash, send_file,session
from modules.db import get_connection
from config import Config
import requests
import shutil
import os
import csv
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from io import BytesIO, StringIO
from flask import Response
from functools import wraps

app = Flask(__name__)
app.config.from_object(Config)

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "usuario" not in session:
            flash("Debes iniciar sesión primero", "danger")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "usuario" not in session:
            flash("Debes iniciar sesión primero", "danger")
            return redirect(url_for("login"))

        if session.get("rol") != "admin":
            flash("No tienes permisos para realizar esta acción", "danger")
            return redirect(url_for("dashboard"))

        return f(*args, **kwargs)
    return decorated_function

def roles_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if "usuario" not in session:
                flash("Debes iniciar sesión primero.", "danger")
                return redirect(url_for("login"))

            if session.get("rol") not in roles:
                flash("No tienes permisos para acceder a esta sección.", "danger")
                return redirect(url_for("dashboard"))

            return f(*args, **kwargs)
        return decorated_function
    return decorator

def get_pagination():
    try:
        page = int(request.args.get("page", 1))
    except ValueError:
        page = 1

    per_page = 10
    offset = (page - 1) * per_page

    return page, per_page, offset

def generar_excel_tabla(nombre_hoja, encabezados, filas, nombre_archivo):
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja

    ws.append(encabezados)

    for fila in filas:
        ws.append(list(fila))

    header_fill = PatternFill(
        start_color="1E293B",
        end_color="1E293B",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    max_row = ws.max_row
    max_col = ws.max_column

    if max_row >= 2:
        tabla_ref = f"A1:{get_column_letter(max_col)}{max_row}"

        tabla = Table(
            displayName=nombre_hoja.replace(" ", "_"),
            ref=tabla_ref
        )

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        tabla.tableStyleInfo = style
        ws.add_table(tabla)

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 4, 60)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form["usuario"].strip()
        password = request.form["password"].strip()

        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT id, nombre, usuario, rol
            FROM usuarios
            WHERE usuario = %s
              AND password = %s
              AND estado = 'ACTIVO';
        """, (usuario, password))

        user = cur.fetchone()

        cur.close()
        conn.close()

        if user:
            session["user_id"] = user[0]
            session["nombre"] = user[1]
            session["usuario"] = user[2]
            session["rol"] = user[3]

            flash(f"Bienvenido, {user[1]}", "success")
            return redirect(url_for("dashboard"))

        flash("Usuario o contraseña incorrectos", "danger")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("Sesión cerrada correctamente", "success")
    return redirect(url_for("login"))

@app.route("/")
def index():
    return redirect(url_for("dashboard"))


@app.route("/dashboard")
@login_required
def dashboard():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM productos;")
    total_productos = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM ventas;")
    total_ventas = cur.fetchone()[0]

    cur.execute("""
            SELECT 
                p.sku,
                p.nombre,
                COALESCE(SUM(v.cantidad), 0) AS total_vendido
            FROM productos p
            LEFT JOIN ventas v ON p.sku = v.sku
            GROUP BY p.sku, p.nombre
            ORDER BY total_vendido DESC;
        """)
    ventas_producto = cur.fetchall()

    cur.execute("""
            SELECT 
                sku,
                nombre,
                stock_actual
            FROM productos
            ORDER BY stock_actual ASC;
        """)
    stock_producto = cur.fetchall()

    cur.execute("""
        SELECT COALESCE(SUM(cantidad), 0)
        FROM ventas
        WHERE DATE(fecha) = CURRENT_DATE;
    """)
    ventas_hoy = cur.fetchone()[0]

    cur.execute("""
        SELECT COUNT(*)
        FROM alertas
        WHERE estado = 'PENDIENTE';
    """)
    total_alertas = cur.fetchone()[0]

    cur.execute("""
        SELECT COUNT(*)
        FROM productos
        WHERE stock_actual <= stock_minimo;
    """)
    productos_criticos = cur.fetchone()[0]

    cur.execute("""
        SELECT sku, nombre, stock_actual, stock_minimo
        FROM productos
        WHERE stock_actual <= stock_minimo
        ORDER BY stock_actual ASC
        LIMIT 5;
    """)
    bajos_stock = cur.fetchall()

    cur.execute("""
        SELECT p.sku, p.nombre, COALESCE(SUM(v.cantidad), 0) AS total_vendido
        FROM productos p
        LEFT JOIN ventas v ON p.sku = v.sku
        GROUP BY p.sku, p.nombre
        ORDER BY total_vendido DESC
        LIMIT 5;
    """)
    top_vendidos = cur.fetchall()

    cur.execute("""
        SELECT sku, COALESCE(SUM(cantidad), 0)
        FROM ventas
        GROUP BY sku
        ORDER BY sku;
    """)
    ventas_productos = cur.fetchall()

    cur.execute("""
        SELECT sku, stock_actual
        FROM productos
        ORDER BY sku;
    """)
    stock_productos = cur.fetchall()

    cur.execute("""
        SELECT id, fecha, sku, cantidad, origen
        FROM ventas
        ORDER BY id DESC
        LIMIT 5;
    """)
    ultimas_ventas = cur.fetchall()

    cur.execute("""
        SELECT id, fecha, sku, tipo, mensaje, estado
        FROM alertas
        WHERE estado = 'PENDIENTE'
        ORDER BY id DESC
        LIMIT 5;
    """)
    ultimas_alertas = cur.fetchall()

    # Ventas por producto
    cur.execute("""
        SELECT 
            p.nombre,
            COALESCE(SUM(v.cantidad), 0) AS total_vendido
        FROM productos p
        LEFT JOIN ventas v ON p.sku = v.sku
        GROUP BY p.nombre
        ORDER BY total_vendido DESC
        LIMIT 8;
    """)
    ventas_producto = cur.fetchall()

    # Stock actual
    cur.execute("""
        SELECT nombre, stock_actual, stock_minimo
        FROM productos
        ORDER BY stock_actual ASC
        LIMIT 8;
    """)
    stock_producto = cur.fetchall()

    # Productos críticos
    cur.execute("""
        SELECT nombre, stock_actual, stock_minimo
        FROM productos
        WHERE stock_actual <= stock_minimo
        ORDER BY stock_actual ASC
        LIMIT 8;
    """)
    productos_bajo_stock = cur.fetchall()

    # Ventas por día
    cur.execute("""
        SELECT 
            DATE(fecha) AS dia,
            COALESCE(SUM(cantidad), 0) AS total
        FROM ventas
        GROUP BY DATE(fecha)
        ORDER BY dia DESC
        LIMIT 7;
    """)
    ventas_dia = cur.fetchall()

    ventas_dia = list(reversed(ventas_dia))

    cur.close()
    conn.close()

    ventas_labels = [v[0] for v in ventas_productos]
    ventas_data = [int(v[1]) for v in ventas_productos]

    stock_labels = [s[0] for s in stock_productos]
    stock_data = [int(s[1]) for s in stock_productos]

    return render_template(
        "dashboard.html",
        total_productos=total_productos,
        total_ventas=total_ventas,
        ventas_hoy=ventas_hoy,
        total_alertas=total_alertas,
        productos_criticos=productos_criticos,
        bajos_stock=bajos_stock,
        top_vendidos=top_vendidos,
        ventas_labels=ventas_labels,
        ventas_data=ventas_data,
        stock_labels=stock_labels,
        stock_data=stock_data,
        ultimas_ventas=ultimas_ventas,
        ultimas_alertas=ultimas_alertas,
        ventas_producto=ventas_producto,
        stock_producto=stock_producto,
        productos_bajo_stock=productos_bajo_stock,
        ventas_dia=ventas_dia
    )


@app.route("/productos")
@roles_required("admin", "inventario", "supervisor")
def productos():
    buscar = request.args.get("buscar", "").strip()
    page, per_page, offset = get_pagination()

    conn = get_connection()
    cur = conn.cursor()

    params = []
    where = ""

    if buscar:
        where = """
            WHERE sku ILIKE %s
               OR nombre ILIKE %s
        """
        params.extend([f"%{buscar}%", f"%{buscar}%"])

    cur.execute(f"""
        SELECT COUNT(*)
        FROM productos
        {where};
    """, params)

    total = cur.fetchone()[0]
    total_pages = max((total + per_page - 1) // per_page, 1)

    cur.execute(f"""
        SELECT id, sku, nombre, stock_actual, stock_minimo, updated_at
        FROM productos
        {where}
        ORDER BY id DESC
        LIMIT %s OFFSET %s;
    """, params + [per_page, offset])

    productos = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "productos.html",
        productos=productos,
        buscar=buscar,
        page=page,
        total_pages=total_pages
    )

@app.route("/productos/agregar", methods=["POST"])
@roles_required("admin", "inventario")
def agregar_producto():
    sku = request.form.get("sku", "").strip().upper()
    nombre = request.form.get("nombre", "").strip()

    try:
        stock_actual = int(request.form.get("stock_actual", 0))
        stock_minimo = int(request.form.get("stock_minimo", 0))
    except ValueError:
        flash("El stock debe ser un número válido.", "danger")
        return redirect(url_for("productos"))

    if not sku or not nombre:
        flash("El SKU y el nombre son obligatorios.", "danger")
        return redirect(url_for("productos"))

    if stock_actual < 0:
        flash("El stock actual no puede ser negativo.", "danger")
        return redirect(url_for("productos"))

    if stock_minimo < 0:
        flash("El stock mínimo no puede ser negativo.", "danger")
        return redirect(url_for("productos"))

    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("SELECT COUNT(*) FROM productos WHERE sku = %s;", (sku,))
        existe = cur.fetchone()[0]

        if existe > 0:
            flash("Ya existe un producto con ese SKU.", "danger")
            cur.close()
            conn.close()
            return redirect(url_for("productos"))

        cur.execute("""
            INSERT INTO productos (sku, nombre, stock_actual, stock_minimo)
            VALUES (%s, %s, %s, %s);
        """, (sku, nombre, stock_actual, stock_minimo))

        cur.execute("""
            INSERT INTO auditoria (evento, detalle)
            VALUES (%s, %s);
        """, (
            "PRODUCTO_CREADO",
            f"Producto agregado: {sku} - {nombre}. Stock inicial: {stock_actual}, mínimo: {stock_minimo}"
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Producto agregado correctamente.", "success")

    except Exception as e:
        flash(f"Error al agregar producto: {str(e)}", "danger")

    return redirect(url_for("productos"))


@app.route("/productos/editar/<int:producto_id>", methods=["POST"])
@roles_required("admin", "inventario")
def editar_producto(producto_id):
    sku = request.form.get("sku", "").strip().upper()
    nombre = request.form.get("nombre", "").strip()

    try:
        stock_actual = int(request.form.get("stock_actual", 0))
        stock_minimo = int(request.form.get("stock_minimo", 0))
    except ValueError:
        flash("El stock debe ser un número válido.", "danger")
        return redirect(url_for("productos"))

    if not sku or not nombre:
        flash("El SKU y el nombre son obligatorios.", "danger")
        return redirect(url_for("productos"))

    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            UPDATE productos
            SET sku = %s,
                nombre = %s,
                stock_actual = %s,
                stock_minimo = %s,
                updated_at = NOW()
            WHERE id = %s;
        """, (sku, nombre, stock_actual, stock_minimo, producto_id))

        cur.execute("""
            INSERT INTO auditoria (evento, detalle)
            VALUES (%s, %s);
        """, (
            "PRODUCTO_EDITADO",
            f"Producto actualizado: {sku} - {nombre}. Stock actual: {stock_actual}, stock mínimo: {stock_minimo}"
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Producto actualizado correctamente.", "success")

    except Exception as e:
        flash(f"Error al actualizar producto: {str(e)}", "danger")

    return redirect(url_for("productos"))


@app.route("/productos/eliminar/<int:producto_id>", methods=["POST"])
@roles_required("admin")
def eliminar_producto(producto_id):
    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT sku, nombre, stock_actual
            FROM productos
            WHERE id = %s;
        """, (producto_id,))

        producto = cur.fetchone()

        if not producto:
            flash("El producto no existe.", "danger")
            cur.close()
            conn.close()
            return redirect(url_for("productos"))

        cur.execute("DELETE FROM productos WHERE id = %s;", (producto_id,))

        cur.execute("""
            INSERT INTO movimientos (sku, tipo, cantidad, detalle, usuario)
            VALUES (%s, %s, %s, %s, %s);
        """, (
            producto[0],
            "ELIMINACION",
            producto[2],
            f"Producto eliminado: {producto[1]}",
            session.get("usuario", "sistema")
        ))

        cur.execute("""
            INSERT INTO auditoria (evento, detalle)
            VALUES (%s, %s);
        """, (
            "PRODUCTO_ELIMINADO",
            f"Producto eliminado: {producto[0]} - {producto[1]}"
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Producto eliminado correctamente.", "success")

    except Exception as e:
        flash(f"Error al eliminar producto: {str(e)}", "danger")

    return redirect(url_for("productos"))


@app.route("/ventas")
@roles_required("admin", "ventas", "supervisor")
def ventas():

    buscar = request.args.get("buscar", "").strip()
    fecha = request.args.get("fecha", "").strip()

    try:
        page = int(request.args.get("page", 1))
    except:
        page = 1

    per_page = 10
    offset = (page - 1) * per_page

    conn = get_connection()
    cur = conn.cursor()

    where = "WHERE 1=1"
    params = []

    if buscar:
        where += " AND sku ILIKE %s"
        params.append(f"%{buscar}%")

    if fecha:
        where += " AND DATE(fecha) = %s"
        params.append(fecha)

    cur.execute(f"""
        SELECT COUNT(*)
        FROM ventas
        {where}
    """, params)

    total = cur.fetchone()[0]
    total_pages = max((total + per_page - 1) // per_page, 1)

    cur.execute(f"""
        SELECT id, fecha, sku, cantidad, origen
        FROM ventas
        {where}
        ORDER BY id DESC
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])

    ventas = cur.fetchall()

    cur.execute("""
        SELECT sku, nombre, stock_actual
        FROM productos
        ORDER BY nombre ASC
    """)

    productos = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "ventas.html",
        ventas=ventas,
        productos=productos,
        buscar=buscar,
        fecha=fecha,
        page=page,
        total_pages=total_pages
    )

@app.route("/ventas/registrar", methods=["POST"])
@roles_required("admin", "ventas")
def registrar_venta():
    sku = request.form.get("sku", "").strip().upper()

    try:
        cantidad = int(request.form.get("cantidad", 0))
    except ValueError:
        flash("La cantidad debe ser un número válido.", "danger")
        return redirect(url_for("ventas"))

    if not sku:
        flash("Debe seleccionar un producto.", "danger")
        return redirect(url_for("ventas"))

    if cantidad <= 0:
        flash("La cantidad debe ser mayor a 0.", "danger")
        return redirect(url_for("ventas"))

    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT stock_actual
            FROM productos
            WHERE sku = %s;
        """, (sku,))

        producto = cur.fetchone()

        cur.close()
        conn.close()

        if not producto:
            flash("El producto seleccionado no existe.", "danger")
            return redirect(url_for("ventas"))

        if cantidad > producto[0]:
            flash("No hay stock suficiente para realizar la venta.", "danger")
            return redirect(url_for("ventas"))

        payload = {
            "sku": sku,
            "cantidad": cantidad,
            "origen": "web"
        }

        response = requests.post(Config.WF2_URL, json=payload, timeout=20)

        try:
            data = response.json()
        except Exception:
            data = {
                "mensaje": response.text or "n8n no devolvió respuesta JSON"
            }

        if response.status_code == 200:
            conn = get_connection()
            cur = conn.cursor()

            cur.execute("""
                INSERT INTO movimientos (sku, tipo, cantidad, detalle, usuario)
                VALUES (%s, %s, %s, %s, %s);
            """, (
                sku,
                "VENTA",
                cantidad,
                "Venta registrada desde la app",
                session.get("usuario", "sistema")
            ))

            conn.commit()
            cur.close()
            conn.close()

            flash(data.get("mensaje", "Venta procesada correctamente."), "success")
        else:
            flash(data.get("mensaje", "Error al procesar venta."), "danger")

    except Exception as e:
        flash(f"Error al conectar con n8n: {str(e)}", "danger")

    return redirect(url_for("ventas"))


@app.route("/alertas")
@roles_required("admin", "inventario", "supervisor")
def alertas():
    buscar = request.args.get("buscar", "").strip()
    estado = request.args.get("estado", "").strip()
    prioridad = request.args.get("prioridad", "").strip()

    conn = get_connection()
    cur = conn.cursor()

    query = """
        SELECT
            id, fecha, sku, tipo, mensaje, estado,
            prioridad, fecha_resolucion, usuario_resolucion
        FROM alertas
        WHERE 1 = 1
    """

    params = []

    if buscar:
        query += " AND (sku ILIKE %s OR tipo ILIKE %s OR mensaje ILIKE %s)"
        params.extend([f"%{buscar}%", f"%{buscar}%", f"%{buscar}%"])

    if estado:
        query += " AND estado = %s"
        params.append(estado)

    if prioridad:
        query += " AND prioridad = %s"
        params.append(prioridad)

    query += """
        ORDER BY
            CASE
                WHEN estado = 'PENDIENTE' THEN 1
                WHEN estado = 'EN_PROCESO' THEN 2
                WHEN estado = 'REABASTECIDA' THEN 3
                WHEN estado = 'ATENDIDA' THEN 4
                ELSE 5
            END,
            id DESC;
    """

    cur.execute(query, params)
    alertas = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "alertas.html",
        alertas=alertas,
        buscar=buscar,
        estado=estado,
        prioridad=prioridad
    )
@app.route("/alertas/estado/<int:alerta_id>/<nuevo_estado>", methods=["POST"])
@roles_required("admin", "inventario")
def cambiar_estado_alerta(alerta_id, nuevo_estado):
    estados_validos = ["PENDIENTE", "EN_PROCESO", "ATENDIDA", "REABASTECIDA"]

    if nuevo_estado not in estados_validos:
        flash("Estado de alerta no válido.", "danger")
        return redirect(url_for("alertas"))

    try:
        conn = get_connection()
        cur = conn.cursor()

        if nuevo_estado in ["ATENDIDA", "REABASTECIDA"]:
            cur.execute("""
                UPDATE alertas
                SET estado = %s,
                    fecha_resolucion = NOW(),
                    usuario_resolucion = %s
                WHERE id = %s;
            """, (
                nuevo_estado,
                session.get("usuario", "sistema"),
                alerta_id
            ))
        else:
            cur.execute("""
                UPDATE alertas
                SET estado = %s
                WHERE id = %s;
            """, (
                nuevo_estado,
                alerta_id
            ))

        cur.execute("""
            INSERT INTO auditoria (evento, detalle)
            VALUES (%s, %s);
        """, (
            "CAMBIO_ESTADO_ALERTA",
            f"Alerta {alerta_id} cambiada a {nuevo_estado} por {session.get('usuario', 'sistema')}"
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Estado de alerta actualizado correctamente.", "success")

    except Exception as e:
        flash(f"Error al actualizar alerta: {str(e)}", "danger")

    return redirect(url_for("alertas"))


@app.route("/reportes")
@roles_required("admin", "supervisor", "inventario")
def reportes():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, fecha, evento, detalle
        FROM auditoria
        WHERE evento = 'REPORTE_SEMANAL_GENERADO'
        ORDER BY id DESC;
    """)
    reportes = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("reportes.html", reportes=reportes)

@app.route("/reportes/generar", methods=["POST"])
@roles_required("admin", "supervisor")
def generar_reporte():

    try:
        response = requests.post(
            Config.WF3_URL,
            json={},
            timeout=30
        )

        try:
            data = response.json()
        except Exception:
            data = {
                "mensaje":
                response.text or "n8n no devolvió respuesta JSON"
            }

        if response.status_code == 200:

            origen = r"C:\Users\edinc\Documents\Cursos_UMG\Ingenieria de software\Proyecto Final inventario_n8n_Inventario_web\inventario-n8n\data\out\reporte_semanal.csv"

            destino = r"C:\Users\edinc\Documents\Cursos_UMG\Ingenieria de software\Proyecto Final inventario_n8n_Inventario_web\inventario_web\reportes_generados\reporte_semanal.csv"

            if os.path.exists(origen):
                shutil.copy(origen, destino)

            flash(
                data.get(
                    "mensaje",
                    "Reporte generado correctamente"
                ),
                "success"
            )

        else:
            flash(
                data.get(
                    "mensaje",
                    "Error al generar reporte"
                ),
                "danger"
            )

    except Exception as e:
        flash(
            f"Error al conectar con n8n: {str(e)}",
            "danger"
        )

    return redirect(url_for("reportes"))


@app.route("/reportes/descargar")
@roles_required("admin", "supervisor")
def descargar_reporte():
    ruta = os.path.join(os.getcwd(), "reporte_semanal.csv")
    if os.path.exists(ruta):
        return send_file(ruta, as_attachment=True)
    flash("No se encontró el archivo del reporte", "danger")
    return redirect(url_for("reportes"))

@app.route("/reportes/descargar-semanal")
@roles_required("admin", "supervisor")
def descargar_reporte_semanal():
    ruta = r"C:\Users\edinc\Documents\Cursos_UMG\Ingenieria de software\Proyecto Final inventario_n8n_Inventario_web\inventario-n8n\data\out\reporte_semanal.csv"

    if not os.path.exists(ruta):
        flash("No existe el reporte semanal. Primero genera el reporte.", "danger")
        return redirect(url_for("reportes"))

    return send_file(
        ruta,
        as_attachment=True,
        download_name="reporte_semanal.csv",
        mimetype="text/csv"
    )

@app.route("/reportes/exportar/ventas-hoy")
@roles_required("admin", "supervisor")
def exportar_ventas_hoy():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, fecha, sku, cantidad, origen
        FROM ventas
        WHERE DATE(fecha) = CURRENT_DATE
        ORDER BY fecha DESC;
    """)

    ventas = cur.fetchall()

    cur.close()
    conn.close()

    return generar_excel_tabla(
        "Ventas Hoy",
        ["ID", "Fecha", "SKU", "Cantidad", "Origen"],
        ventas,
        "ventas_hoy.xlsx"
    )

@app.route("/reportes/exportar/ventas-semana")
@roles_required("admin", "supervisor")
def exportar_ventas_semana():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, fecha, sku, cantidad, origen
        FROM ventas
        WHERE fecha >= NOW() - INTERVAL '7 days'
        ORDER BY fecha DESC;
    """)

    ventas = cur.fetchall()

    cur.close()
    conn.close()

    return generar_excel_tabla(
        "Ventas Semana",
        ["ID", "Fecha", "SKU", "Cantidad", "Origen"],
        ventas,
        "ventas_semana.xlsx"
    )

@app.route("/reportes/exportar/alertas")
@roles_required("admin", "supervisor")
def exportar_alertas():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, fecha, sku, tipo, mensaje, estado, prioridad
        FROM alertas
        WHERE estado IN ('PENDIENTE', 'EN_PROCESO')
        ORDER BY fecha DESC;
    """)

    alertas = cur.fetchall()

    cur.close()
    conn.close()

    return generar_excel_tabla(
        "Alertas",
        ["ID", "Fecha", "SKU", "Tipo", "Mensaje", "Estado", "Prioridad"],
        alertas,
        "alertas_pendientes.xlsx"
    )


@app.route("/reportes/exportar/inventario-critico")
@roles_required("admin", "supervisor")
def exportar_inventario_critico():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT sku, nombre, stock_actual, stock_minimo
        FROM productos
        WHERE stock_actual <= stock_minimo
        ORDER BY stock_actual ASC;
    """)

    productos = cur.fetchall()

    cur.close()
    conn.close()

    return generar_excel_tabla(
        "Inventario Critico",
        ["SKU", "Nombre", "Stock Actual", "Stock Mínimo"],
        productos,
        "inventario_critico.xlsx"
    )

@app.route("/reabastecimiento", methods=["GET", "POST"])
@roles_required("admin", "inventario")
def reabastecimiento():
    conn = get_connection()
    cur = conn.cursor()

    if request.method == "POST":
        sku = request.form.get("sku", "").strip().upper()
        motivo = request.form.get("motivo", "").strip()
        usuario = session.get("usuario", "sistema")

        try:
            cantidad = int(request.form.get("cantidad", 0))
        except ValueError:
            flash("La cantidad debe ser un número válido.", "danger")
            return redirect(url_for("reabastecimiento"))

        if not sku:
            flash("Debe seleccionar un producto.", "danger")
            return redirect(url_for("reabastecimiento"))

        if cantidad <= 0:
            flash("La cantidad debe ser mayor a 0.", "danger")
            return redirect(url_for("reabastecimiento"))

        if not motivo:
            flash("Debe ingresar un motivo del reabastecimiento.", "danger")
            return redirect(url_for("reabastecimiento"))

        try:
            cur.execute("""
                SELECT COUNT(*)
                FROM productos
                WHERE sku = %s;
            """, (sku,))

            existe = cur.fetchone()[0]

            if existe == 0:
                flash("El producto seleccionado no existe.", "danger")
                cur.close()
                conn.close()
                return redirect(url_for("reabastecimiento"))

            cur.execute("""
                UPDATE productos
                SET stock_actual = stock_actual + %s,
                    updated_at = NOW()
                WHERE sku = %s;
            """, (cantidad, sku))

            cur.execute("""
                INSERT INTO reabastecimientos (sku, cantidad, motivo, usuario)
                VALUES (%s, %s, %s, %s);
            """, (sku, cantidad, motivo, usuario))

            cur.execute("""
                INSERT INTO movimientos (sku, tipo, cantidad, detalle, usuario)
                VALUES (%s, %s, %s, %s, %s);
            """, (
                sku,
                "REABASTECIMIENTO",
                cantidad,
                motivo,
                usuario
            ))

            cur.execute("""
                UPDATE alertas
                SET estado = 'REABASTECIDA',
                    fecha_resolucion = NOW(),
                    usuario_resolucion = %s
                WHERE sku = %s
                AND tipo = 'BAJO_STOCK'
                AND estado IN ('PENDIENTE', 'EN_PROCESO')
                AND (
                    SELECT stock_actual FROM productos WHERE sku = %s
                ) > (
                    SELECT stock_minimo FROM productos WHERE sku = %s
                );
            """, (usuario, sku, sku, sku))

            cur.execute("""
                INSERT INTO auditoria (evento, detalle)
                VALUES (%s, %s);
            """, (
                "REABASTECIMIENTO",
                f"Se agregaron {cantidad} unidades al producto {sku}. Motivo: {motivo}"
            ))

            conn.commit()
            flash("Reabastecimiento registrado correctamente.", "success")

        except Exception as e:
            conn.rollback()
            flash(f"Error al registrar reabastecimiento: {str(e)}", "danger")

    cur.execute("""
        SELECT sku, nombre, stock_actual, stock_minimo
        FROM productos
        ORDER BY nombre;
    """)
    productos = cur.fetchall()

    cur.execute("""
        SELECT id, fecha, sku, cantidad, motivo, usuario
        FROM reabastecimientos
        ORDER BY id DESC
        LIMIT 20;
    """)
    historial = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "reabastecimiento.html",
        productos=productos,
        historial=historial
    )

@app.route("/movimientos")
@roles_required("admin", "inventario", "supervisor")
def movimientos():
    buscar = request.args.get("buscar", "").strip()
    tipo = request.args.get("tipo", "").strip()
    usuario = request.args.get("usuario", "").strip()

    conn = get_connection()
    cur = conn.cursor()

    query = """
        SELECT id, fecha, sku, tipo, cantidad, detalle, usuario
        FROM movimientos
        WHERE 1 = 1
    """

    params = []

    if buscar:
        query += " AND (sku ILIKE %s OR detalle ILIKE %s)"
        params.extend([f"%{buscar}%", f"%{buscar}%"])

    if tipo:
        query += " AND tipo = %s"
        params.append(tipo)

    if usuario:
        query += " AND usuario ILIKE %s"
        params.append(f"%{usuario}%")

    query += " ORDER BY id DESC LIMIT 100;"

    cur.execute(query, params)
    movimientos = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "movimientos.html",
        movimientos=movimientos,
        buscar=buscar,
        tipo=tipo,
        usuario=usuario
    )

@app.route("/auditoria")
@roles_required("admin", "supervisor")
def auditoria():
    buscar = request.args.get("buscar", "").strip()

    conn = get_connection()
    cur = conn.cursor()

    if buscar:
        cur.execute("""
            SELECT id, fecha, evento, detalle
            FROM auditoria
            WHERE evento ILIKE %s
               OR detalle ILIKE %s
            ORDER BY id DESC;
        """, (
            f"%{buscar}%",
            f"%{buscar}%"
        ))
    else:
        cur.execute("""
            SELECT id, fecha, evento, detalle
            FROM auditoria
            ORDER BY id DESC;
        """)

    auditorias = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "auditoria.html",
        auditorias=auditorias,
        buscar=buscar
    )

@app.route("/reportes/inventario/excel")
@roles_required("admin", "supervisor")
def descargar_inventario_excel():

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            sku,
            nombre,
            stock_actual,
            stock_minimo,
            updated_at
        FROM productos
        ORDER BY sku;
    """)

    productos = cur.fetchall()

    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    encabezados = [
        "SKU",
        "Nombre",
        "Stock Actual",
        "Stock Mínimo",
        "Última Actualización"
    ]

    ws.append(encabezados)

    fill = PatternFill(
        start_color="1E293B",
        end_color="1E293B",
        fill_type="solid"
    )

    font = Font(
        color="FFFFFF",
        bold=True
    )

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font

    for p in productos:
        ws.append(p)

    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 5

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return Response(
        excel_file.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            "attachment; filename=reporte_inventario.xlsx"
        }
    )

@app.route("/reportes/ventas/excel")
@roles_required("admin", "supervisor")
def descargar_ventas_excel():

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            id,
            fecha,
            sku,
            cantidad,
            origen
        FROM ventas
        ORDER BY id DESC;
    """)

    ventas = cur.fetchall()

    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    encabezados = [
        "ID",
        "Fecha",
        "SKU",
        "Cantidad",
        "Origen"
    ]

    ws.append(encabezados)

    fill = PatternFill(
        start_color="2563EB",
        end_color="2563EB",
        fill_type="solid"
    )

    font = Font(
        color="FFFFFF",
        bold=True
    )

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font

    for v in ventas:
        ws.append(v)

    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 5

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return Response(
        excel_file.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            "attachment; filename=reporte_ventas.xlsx"
        }
    )

@app.route("/reportes/alertas/excel")
@roles_required("admin", "supervisor")
def descargar_alertas_excel():

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            id,
            fecha,
            sku,
            tipo,
            mensaje,
            estado
        FROM alertas
        ORDER BY id DESC;
    """)

    alertas = cur.fetchall()

    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Alertas"

    encabezados = [
        "ID",
        "Fecha",
        "SKU",
        "Tipo",
        "Mensaje",
        "Estado"
    ]

    ws.append(encabezados)

    fill = PatternFill(
        start_color="DC2626",
        end_color="DC2626",
        fill_type="solid"
    )

    font = Font(
        color="FFFFFF",
        bold=True
    )

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font

    for a in alertas:
        ws.append(a)

    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 5

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return Response(
        excel_file.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            "attachment; filename=reporte_alertas.xlsx"
        }
    )

@app.route("/configuracion", methods=["GET", "POST"])
@roles_required("admin")
def configuracion():
    conn = get_connection()
    cur = conn.cursor()

    if request.method == "POST":
        nombre_negocio = request.form.get("nombre_negocio", "").strip()
        correo_alertas = request.form.get("correo_alertas", "").strip()

        try:
            stock_minimo_global = int(request.form.get("stock_minimo_global", 5))
        except ValueError:
            flash("El stock mínimo global debe ser un número válido.", "danger")
            return redirect(url_for("configuracion"))

        if not nombre_negocio:
            flash("El nombre del negocio es obligatorio.", "danger")
            return redirect(url_for("configuracion"))

        if stock_minimo_global < 0:
            flash("El stock mínimo global no puede ser negativo.", "danger")
            return redirect(url_for("configuracion"))

        logo_path = None
        logo_file = request.files.get("logo")

        if logo_file and logo_file.filename:
            filename = secure_filename(logo_file.filename)
            extension = filename.rsplit(".", 1)[-1].lower()

            if extension not in ["png", "jpg", "jpeg", "webp"]:
                flash("El logo debe ser PNG, JPG, JPEG o WEBP.", "danger")
                return redirect(url_for("configuracion"))

            nuevo_nombre = f"logo_sistema.{extension}"
            carpeta_logo = os.path.join("static", "img", "logos")
            os.makedirs(carpeta_logo, exist_ok=True)

            ruta_logo = os.path.join(carpeta_logo, nuevo_nombre)
            logo_file.save(ruta_logo)

            logo_path = f"img/logos/{nuevo_nombre}"

        try:
            if logo_path:
                cur.execute("""
                    UPDATE configuracion_sistema
                    SET nombre_negocio = %s,
                        correo_alertas = %s,
                        stock_minimo_global = %s,
                        logo = %s,
                        updated_at = NOW()
                    WHERE id = (
                        SELECT id FROM configuracion_sistema
                        ORDER BY id ASC
                        LIMIT 1
                    );
                """, (
                    nombre_negocio,
                    correo_alertas,
                    stock_minimo_global,
                    logo_path
                ))
            else:
                cur.execute("""
                    UPDATE configuracion_sistema
                    SET nombre_negocio = %s,
                        correo_alertas = %s,
                        stock_minimo_global = %s,
                        updated_at = NOW()
                    WHERE id = (
                        SELECT id FROM configuracion_sistema
                        ORDER BY id ASC
                        LIMIT 1
                    );
                """, (
                    nombre_negocio,
                    correo_alertas,
                    stock_minimo_global
                ))

            cur.execute("""
                INSERT INTO auditoria (evento, detalle)
                VALUES (%s, %s);
            """, (
                "CONFIGURACION_ACTUALIZADA",
                f"Configuración del sistema actualizada por {session.get('usuario', 'sistema')}"
            ))

            conn.commit()
            flash("Configuración actualizada correctamente.", "success")

        except Exception as e:
            conn.rollback()
            flash(f"Error al actualizar configuración: {str(e)}", "danger")

    cur.execute("""
        SELECT id, nombre_negocio, correo_alertas, stock_minimo_global, logo, updated_at
        FROM configuracion_sistema
        ORDER BY id ASC
        LIMIT 1;
    """)

    config = cur.fetchone()

    cur.close()
    conn.close()

    return render_template("configuracion.html", config=config)

@app.route("/usuarios")
@roles_required("admin")
def usuarios():

    buscar = request.args.get("buscar", "").strip()

    conn = get_connection()
    cur = conn.cursor()

    if buscar:

        cur.execute("""
            SELECT
                id,
                nombre,
                usuario,
                rol,
                created_at
            FROM usuarios
            WHERE nombre ILIKE %s
               OR usuario ILIKE %s
               OR rol ILIKE %s
            ORDER BY id DESC;
        """, (
            f"%{buscar}%",
            f"%{buscar}%",
            f"%{buscar}%"
        ))

    else:

        cur.execute("""
            SELECT
                id,
                nombre,
                usuario,
                rol,
                created_at
            FROM usuarios
            ORDER BY id DESC;
        """)

    usuarios = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "usuarios.html",
        usuarios=usuarios,
        buscar=buscar
    )

@app.route("/usuarios/agregar", methods=["POST"])
@roles_required("admin")
def agregar_usuario():

    nombre = request.form.get("nombre", "").strip()
    usuario = request.form.get("usuario", "").strip()
    password = request.form.get("password", "").strip()
    rol = request.form.get("rol", "").strip()

    roles_validos = [
        "admin",
        "inventario",
        "ventas",
        "supervisor"
    ]

    if not nombre or not usuario or not password:

        flash("Todos los campos son obligatorios.", "danger")
        return redirect(url_for("usuarios"))

    if rol not in roles_validos:

        flash("Rol inválido.", "danger")
        return redirect(url_for("usuarios"))

    try:

        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT COUNT(*)
            FROM usuarios
            WHERE usuario = %s;
        """, (usuario,))

        existe = cur.fetchone()[0]

        if existe > 0:

            flash("Ya existe un usuario con ese nombre.", "danger")

            cur.close()
            conn.close()

            return redirect(url_for("usuarios"))

        cur.execute("""
            INSERT INTO usuarios (
                nombre,
                usuario,
                password,
                rol
            )
            VALUES (%s, %s, %s, %s);
        """, (
            nombre,
            usuario,
            password,
            rol
        ))

        cur.execute("""
            INSERT INTO auditoria (evento, detalle)
            VALUES (%s, %s);
        """, (
            "USUARIO_CREADO",
            f"Usuario creado: {usuario} ({rol})"
        ))

        conn.commit()

        cur.close()
        conn.close()

        flash("Usuario agregado correctamente.", "success")

    except Exception as e:

        flash(f"Error al agregar usuario: {str(e)}", "danger")

    return redirect(url_for("usuarios"))

@app.route("/usuarios/editar/<int:usuario_id>", methods=["POST"])
@roles_required("admin")
def editar_usuario(usuario_id):

    nombre = request.form.get("nombre", "").strip()
    usuario = request.form.get("usuario", "").strip()
    rol = request.form.get("rol", "").strip()

    roles_validos = [
        "admin",
        "inventario",
        "ventas",
        "supervisor"
    ]

    if not nombre or not usuario:

        flash("Todos los campos son obligatorios.", "danger")
        return redirect(url_for("usuarios"))

    if rol not in roles_validos:

        flash("Rol inválido.", "danger")
        return redirect(url_for("usuarios"))

    try:

        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT COUNT(*)
            FROM usuarios
            WHERE usuario = %s
              AND id <> %s;
        """, (
            usuario,
            usuario_id
        ))

        existe = cur.fetchone()[0]

        if existe > 0:

            flash("Ya existe otro usuario con ese nombre.", "danger")

            cur.close()
            conn.close()

            return redirect(url_for("usuarios"))

        cur.execute("""
            UPDATE usuarios
            SET nombre = %s,
                usuario = %s,
                rol = %s
            WHERE id = %s;
        """, (
            nombre,
            usuario,
            rol,
            usuario_id
        ))

        cur.execute("""
            INSERT INTO auditoria (evento, detalle)
            VALUES (%s, %s);
        """, (
            "USUARIO_EDITADO",
            f"Usuario actualizado: {usuario}"
        ))

        conn.commit()

        cur.close()
        conn.close()

        flash("Usuario actualizado correctamente.", "success")

    except Exception as e:

        flash(f"Error al actualizar usuario: {str(e)}", "danger")

    return redirect(url_for("usuarios"))

@app.route("/usuarios/eliminar/<int:usuario_id>", methods=["POST"])
@roles_required("admin")
def eliminar_usuario(usuario_id):

    try:

        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT usuario
            FROM usuarios
            WHERE id = %s;
        """, (usuario_id,))

        usuario = cur.fetchone()

        if not usuario:

            flash("Usuario no encontrado.", "danger")

            cur.close()
            conn.close()

            return redirect(url_for("usuarios"))

        cur.execute("""
            DELETE FROM usuarios
            WHERE id = %s;
        """, (usuario_id,))

        cur.execute("""
            INSERT INTO auditoria (evento, detalle)
            VALUES (%s, %s);
        """, (
            "USUARIO_ELIMINADO",
            f"Usuario eliminado: {usuario[0]}"
        ))

        conn.commit()

        cur.close()
        conn.close()

        flash("Usuario eliminado correctamente.", "success")

    except Exception as e:

        flash(f"Error al eliminar usuario: {str(e)}", "danger")

    return redirect(url_for("usuarios"))

@app.route("/reportes/exportar/auditoria")
@roles_required("admin", "supervisor")
def exportar_auditoria_excel():
    buscar = request.args.get("buscar", "").strip()
    evento = request.args.get("evento", "").strip()
    fecha = request.args.get("fecha", "").strip()

    conn = get_connection()
    cur = conn.cursor()

    where = "WHERE 1 = 1"
    params = []

    if buscar:
        where += " AND detalle ILIKE %s"
        params.append(f"%{buscar}%")

    if evento:
        where += " AND evento ILIKE %s"
        params.append(f"%{evento}%")

    if fecha:
        where += " AND DATE(fecha) = %s"
        params.append(fecha)

    cur.execute(f"""
        SELECT id, fecha, evento, detalle
        FROM auditoria
        {where}
        ORDER BY id DESC;
    """, params)

    auditorias = cur.fetchall()

    cur.close()
    conn.close()

    return generar_excel_tabla(
        "Auditoria",
        ["ID", "Fecha", "Evento", "Detalle"],
        auditorias,
        "reporte_auditoria.xlsx"
    )

@app.context_processor
def cargar_notificaciones():
    try:
        if "usuario" not in session:
            return {}

        conn = get_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT COUNT(*)
            FROM alertas
            WHERE estado IN ('PENDIENTE', 'EN_PROCESO');
        """)
        total_notificaciones = cur.fetchone()[0]

        cur.execute("""
            SELECT id, sku, mensaje, prioridad, estado
            FROM alertas
            WHERE estado IN ('PENDIENTE', 'EN_PROCESO')
            ORDER BY id DESC
            LIMIT 5;
        """)
        notificaciones = cur.fetchall()

        cur.close()
        conn.close()

        return {
            "total_notificaciones": total_notificaciones,
            "notificaciones": notificaciones
        }

    except Exception:
        return {
            "total_notificaciones": 0,
            "notificaciones": []
        }

if __name__ == "__main__":
    app.run(debug=True, port=5000)


